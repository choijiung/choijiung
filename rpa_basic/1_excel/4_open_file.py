from openpyxl import load_workbook # 파일 불러오기
wb = load_workbook("sample_cell.xlsx") # sample_cell 파일에서 wb을 불러옴
ws = wb.active # 활성화된 Sheet

# cell 데이터 불러오기
for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(row=x, column=y).value, end=" ") # 1, 2, 3, 4..
    print()
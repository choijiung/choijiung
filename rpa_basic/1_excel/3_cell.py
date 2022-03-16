from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "cell sheet"

ws["A1"] = 1 # A1 셀이 1이라는 값을 입력
ws["A2"] = 2
ws["A3"] = 3

ws["b1"] = 4
ws["b2"] = 5
ws["b3"] = 6

print(ws["A1"])# A1의 셀 정보를 출력
print(ws["A1"].value) # A1 셀의 '값 출력
print(ws["A10"].value) # 값이 없을땐 none 출력

#row = 1, 2, 3 ...
#column = A, B, C ...
print(ws.cell(column=1,row=1).value) # ws["A1"].value
print(ws.cell (column=2, row=1).value) # ws["B1"].value

c = ws.cell(column=3, row=1, value=10) # ws["C1"].value = 10
print(c.value) # ws["C1"].value

from random import *


# 반복문을 이용하여 랜덤 숫자 채우기
index = 1
for x in range(1, 11):
    for y in range(1, 11):
        # ws.cell(row=x, column=y, value=randint(0, 100)) # 0~100 의 숫자
        ws.cell(row=x, column=y, value=index)
        index += 1

wb.save("excel_files/sample_cell.xlsx")
wb.close()
from openpyxl import Workbook
wb = Workbook()
#ws = wb.active
ws = wb.create_sheet() # 새로운 Sheet 생성
ws.title = "MySheet"
ws.sheet_properties.tabColor = "00FFFF" # RGB 형태로 색을 넣으면 색상 변경


# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 시트 생성
ws2 = wb.create_sheet("NEW Sheet", 2) #2번째 자리에 Sheet 생성 (코딩의 새는 순서는 앞에서부터 0, 1, 2..)

#ws2.명령어
new_ws = wb["NEW Sheet"] # 딕셔너리 형태로 시트에 접근

print(wb.sheetnames) # 모든 Sheet 이름 출력

#Sheet 복사
new_ws["A1"] = "TEST" # 엑셀의 A1셀에 TEST 입력
target = wb.copy_worksheet(new_ws) #new_ws Sheet 의 값을 복사해서 새 Sheet에 저장
target.title = "Copied Sheet" # 새 Sheet의 이름을 Copied Sheet 으로 저장

wb.save("excel_files/sample_sheet.xlsx")
wb.close()